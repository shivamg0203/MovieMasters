
from distutils.log import error
from functools import wraps
import io
import json
from math import floor
from sqlalchemy import or_, and_
from pickle import NONE
import pandas as pd
import sqlite3
from datetime import time
from sre_constants import SUCCESS
from unicodedata import name
from flask import Blueprint, abort, render_template,  flash, jsonify, redirect, url_for, send_file,make_response
from flask_login import *
from .models import User, Movie,Theatre,Venue,Order
from . import db
from .auth import *
from datetime import *
from jsons import JsonSerializable,JsonsError
from flask import sessions,session
import base64
from jinja2 import Environment, environment,filters
import openpyxl


views = Blueprint('views', __name__)
 
# JINJA TEMPLATE FILTERS

@views.app_template_filter('base64_encode')    # jinja template filter base64_encode
def base64_encode(val):
    return base64.b64encode(val).decode('utf-8')

@views.app_template_filter('seatsbook')    # jinja template filter seatsbook
def seatsbook(val):
    return seatsbooked(val)
 
# FUNCTIONS FOR CALCULATIONS 

def seatsbooked(venueid):
    seat=0
    allorder=Order.query.all()
    for q in allorder:
        if (q.venue_id==int(venueid)):
            seat=seat+q.seats
    return seat 
    
def returntheatre(city):
    session['city']=city
    theatres=Theatre.query.filter_by(city=city).all()
    return theatres

def returnmovies(city):
    movies=[]
    session['city']=city
    for i in Venue.query.all():
        if(i.theatre_venue.city==city and i.date>=date.today()):
            if i.movie_venue not in movies:
                movies.append(i.movie_venue)
    return movies    

def booktickets(venue_id,tickets):
    reqvenue=Venue.query.get(venue_id)
    new_order=Order(venue_id=venue_id, user_id=current_user.id,seats=tickets)
    db.session.add(new_order)
    db.session.commit() 
    session['orderid']=new_order.id
    return session['orderid']  
     
def returnallorders(uid):
    orderlist=[]
    orderlist=Order.query.filter_by(user_id=uid).all()
    return orderlist

def venuebytheatre(tid):
    venuelist=[]
    venuelist=Venue.query.filter(Venue.theatre_id==tid, Venue.date>=date.today()).all()
    return venuelist

def venuebymovie(mid):
    venuelist=[]

    for k in Venue.query.all():
        if(mid==k.movie_id and k.theatre_venue.city==session['city'] and k.date>=date.today()):
            venuelist.append(k)  
    return venuelist

def updateimage(a,b):
    if (a.filename==''):
        print(True)
    else:
        image=request.files['image']
        image.save('website/static/new.png')
        with open('website/static/new.png', 'rb') as imgfile:
            photo= imgfile.read() 
            b.image=photo  
  
def allvenue():
    return Venue.query.all()

def alltheatre():
    return Theatre.query.all()

def allmovie():
    return Movie.query.all()

def returndate(d):
    return date(int((d)[0:4]),int((d)[5:7]),int((d)[8:]))

def returntime(t):
    return time(int((t)[:2]),int((t)[3:5]))
           
def userobject(id):
    return User.query.get(id)  

def admin_required(f):      # Adding admin privilages
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.username != 'shivam':
            return abort(401)
        return f(*args, **kwargs)
    return decorated_function 

    
    

@views.route('/home/', methods=['GET', 'POST'])               #route for homepage
@login_required
def home():        #homefunction takes username as argument 
    if 'city' not in session:
        session['city']='Mumbai'
    moviecount= Movie.query.count()
    recomm=Movie.query.filter(Movie.id <= floor(moviecount/2)).all()
    pre=Movie.query.filter(and_(Movie.id > floor(moviecount/2) , Movie.id <= floor(moviecount*0.6))).all()
    other=Movie.query.filter(and_(Movie.id > floor(moviecount*0.6) ,Movie.id <= moviecount)).all()   
    if request.method=='POST' and 'setcity' in request.form:
        city=request.form.get('setcity')
        session['city']=city
        return render_template("home.html",city=city,user=current_user, recomm=recomm,pre=pre,other=other,movies=returnmovies(city),theatres=returntheatre(city))   
    if request.method=='GET':

        return render_template("home.html",city=session['city'],user=current_user, recomm=recomm,pre=pre,other=other,movies=returnmovies(session['city']),theatres=returntheatre(session['city']))

@views.route('/searchbytheatre', methods=['GET', 'POST'])              
@login_required
def searchbytheatre():
    if request.method=='POST' and 'theatresearch' in request.form :
        theatre=request.form.get('theatredat')  #theatre = Theatre_Name (City_Name)
        l=len(session['city'])
        name=theatre[0:len(theatre)-l-3]   # Store name of theatre in 'name' 
        theatrename=Theatre.query.filter_by(name=name, city=session['city']).first()
        if theatrename==None:
            flash('Theatre does not exist')
            return redirect(url_for('views.home'))  
        venuelist=venuebytheatre(theatrename.id)         
        if venuelist==[]:
            flash('No venues of '+theatrename.name+' in your city')
            return redirect(url_for('views.home')) 
        session['theatreid']=theatrename.id    # May Not Be Needed
        return render_template('searchbytheatre.html',  venuelist=venuelist,theatrename=theatrename,movies=returnmovies(session['city']),theatres=returntheatre(session['city']))
    
    if request.method=='POST' and 'setcity' in request.form:  
        session['city']=request.form.get('setcity')
        return redirect(url_for('views.home')) 
    
    elif request.method=='POST' and 'tbookvenue' in request.form:
        rid=request.form.get('tbookvenue')
        if rid==None:
            flash("Invalid Venue")
            return redirect(url_for('views.searchbytheatre'))
        tickets=request.form.get('numberoftickets-'+str(rid)) #GET CORRESPONDING NUMBER OF TICKETS
        hallsize=(Venue.query.get(rid).theatre_venue.capacity)-seatsbooked(rid)
        if (tickets.isdigit()==False or int(tickets)<0 or int(tickets)>hallsize ):
            flash("Invalid Number of Tickets")
            return redirect(url_for('views.searchbytheatre'))
        oid=booktickets(int(rid),int(tickets))
        return redirect(url_for('views.book', venue_id=int(rid),tickets=int(tickets),oid=oid)) 
          
    if request.method=='GET':
        venuelist=venuebytheatre(session.get('theatreid')) 
        return render_template('searchbytheatre.html', venuelist=venuelist,theatrename=theatrename,movies=returnmovies(session['city']),theatres=returntheatre(session['city']))

@views.route('/searchbymovie', methods=['GET', 'POST'])              
@login_required
def searchbymovie(): 
    if request.method=='POST' and 'moviesearch' in request.form :
        moviename=request.form.get('moviedata') 
        moviedata=Movie.query.filter_by(name=moviename).first()
        if moviedata==None:
            flash('Movie does not exist')
            return redirect(url_for('views.home')) 
        session['movieid']=moviedata.id
        venuelist=venuebymovie(moviedata.id)
        
        if venuelist==[]:
            flash('No venues of '+moviename+' in your city')
            return redirect(url_for('views.home')) 
        return render_template('searchbymovie.html', venuelist=venuelist,moviedata=moviedata,movies=returnmovies(session['city']),theatres=returntheatre(session['city']))
    
    elif request.method=='POST' and 'mbookvenue' in request.form:
        rid=request.form.get('mbookvenue')
        if rid==None:
            flash("Invalid Venue")
            return redirect(url_for('views.searchbymovie'))
        tickets=request.form.get('number_of_tickets-'+str(rid)) #GET CORRESPONDING NUMBER OF TICKETS
        hallsize=(Venue.query.get(rid).theatre_venue.capacity)-seatsbooked(rid)
        if (tickets.isdigit()==False or int(tickets)<0 or int(tickets)>hallsize):
            flash("Invalid Number of Tickets")
            return redirect(url_for('views.searchbymovie'))
        
        oid=booktickets(int(rid),int(tickets))
        return redirect(url_for('views.book', venue_id=int(rid),tickets=int(tickets),oid=oid)) 
        
    elif request.method=='GET':
        if (session.get('movieid')==None):
            return abort(404)
        moviedata=Movie.query.get(session['movieid'])
        venuelist=venuebymovie(session['movieid'])
        if venuelist==[]:
            flash('No venues of '+moviedata.name+' in your city')
            return redirect(url_for('views.home')) 
        return render_template('searchbymovie.html', venuelist=venuelist,moviedata=moviedata,movies=returnmovies(session['city']),theatres=returntheatre(session['city']))
    elif request.method=='POST' and 'setcity' in request.form: 
        session['city']=request.form.get('setcity')
        return redirect(url_for('views.home')) 
       

@views.route('/searchbymovie/<moviename>', methods=['GET', 'POST'])              
@login_required
def searchbymoviename(moviename): 
    moviedata=Movie.query.filter_by(name=moviename).first()
    session['movieid']=moviedata.id
    return redirect(url_for('views.searchbymovie'))
    
    
@views.route('book/<int:venue_id>/<int:tickets>/<int:oid>', methods=['GET', 'POST'])              
@login_required
def book(venue_id,tickets,oid): 
    reqvenue=Venue.query.get(venue_id)
    if request.method=='GET':
        return render_template('book.html', reqvenue=reqvenue,tickets=tickets,oid=oid) 
    elif request.method=='POST' and 'setcity' in request.form: 
        session['city']=request.form.get('setcity')
        return redirect(url_for('views.home'))    
    return render_template('book.html', reqvenue=reqvenue,tickets=tickets,oid=oid)     

@views.route('/myorders', methods=['GET', 'POST'])              
@login_required
def myorders():   
    if request.method=='POST' and 'setcity' in request.form: 
        session['city']=request.form.get('setcity')
        return redirect(url_for('views.home'))    
  
    orderlist = returnallorders(current_user.id)
    if orderlist==[]:
        flash("You don't have any orders")
        return redirect(url_for('views.home')) 
    if request.method=='GET':     
        return render_template('order.html', orderlist=orderlist, today=date.today())
    
@views.route('/order_in_excel', methods=['GET', 'POST'])      #User can export or download his orders in excel file.
@login_required      
def order_in_excel():
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    # Select the active worksheet
    worksheet = workbook.active
    # Define the headers for the columns
    headers = ['Movie Name', 'Theatre Name', 'City', 'Venue Date', 'Venue Time', 'Seats','Cost','Total Amount']
    # Write the headers to the first row of the worksheet  
    orders=[]
    for order in returnallorders(current_user.id):
        orders.append({'Movie_Name': order.order_venue.movie_venue.name,
                    'Theatre_Name': order.order_venue.theatre_venue.name,
                    'City': order.order_venue.theatre_venue.city,
                    'Venue_Date': order.order_venue.date,
                    'Venue_Time':order.order_venue.time,
                    'Seats':order.seats,
                    'Cost':order.order_venue.cost,
                    'Total_Amount': order.seats * order.order_venue.cost})
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    # Select the active worksheet
    worksheet = workbook.active
    # Write the headers to the first row of the worksheet
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
    # Write each order object to a row in the worksheet
    for row_num, order in enumerate(orders, 2):
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = str(order[header.replace(' ', '_')])
            # cell.value = order[str(header)]
    # Save the workbook to a byte stream
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    # Send the byte stream as a downloadable attachment
    return send_file(file_stream, download_name=f'orders_{current_user.username}.xlsx', as_attachment=True)
    
@views.route('/download_in_excel/<string:name>', methods=['GET', 'POST'])   #Admin can export or download Movie/Theatre/Venue in excel file.
@admin_required
@login_required      
def download_in_excel(name):
    req_list=[]
    if name=='venue':
        for order in allvenue():
            req_list.append({'Id':order.id,
                        'Theatre Id':order.theatre_id,
                        'Movie Id':order.movie_id,
                        'Venue Date':order.date,
                        'Venue Time':order.time,
                        'Cost':order.cost,})
        headers=['Id', 'Theatre Id', 'Movie Id', 'Venue Date','Venue Time','Cost' ]
        
    elif name=='movie':
        for order in allmovie():
            req_list.append({'Id':order.id,
                        'Name':order.name,
                        'Actors':str(order.actors),
                        'Duration': order.duration,
                        'About':order.about,
                        'IMDB':order.imdb_rating,
                        'Released':order.release_date,
                        'Director':order.director_name,
                        'Feedback':'Good'})
        headers=['Id', 'Name', 'Actors', 'Duration','About','IMDB','Released','Director','Feedback']
        
    elif name=='theatre':
        for order in alltheatre():
            req_list.append({'Id': order.id,
                        'Theatre Name': order.name,
                        'Total Capacity': order.capacity,
                        'City':order.city})
        headers=['Id', 'Theatre Name', 'Total Capacity', 'City']
    else:
        flash('Unknown URL')
        return redirect(url_for('views.admin'))
    if req_list==[]:
        flash('No Data')
        return redirect(url_for('views..admin', movies=allmovie))
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
    for row_num, k in enumerate(req_list, 2):
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = k[header]    
            
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    # Send the byte stream as a downloadable attachment
    return send_file(file_stream, download_name=f'data_in_xl_{name}.xlsx', as_attachment=True)    
    



    
@views.route('/admin', methods=['GET', 'POST'])  # URL endpoint for movies interface where user can edit, add and delete movies 
@admin_required                                  # depending upon conditions and Venue constraints.                                   
@login_required
def admin(): 

    if request.method=='GET':
        return render_template("admin.html", movies=allmovie())
    
    elif request.method=='POST':
        if 'edit' in request.form :
            i=request.form['edit']
            toedit=Movie.query.get(i)
            if toedit==None:
                flash("Movie does not exist")
                return render_template("admin.html",movies=allmovie()) 
            return render_template("editmovie.html",p=toedit) 
        if 'moviesearch' in request.form :
            i=request.form.get('editme')
            toedit=Movie.query.get(i)
            if toedit==None:
                flash("Movie does not exist")
                return render_template("admin.html",movies=allmovie()) 
            return render_template("editmovie.html",p=toedit) 
        elif 'editmovie' in request.form:
            editmovie=request.form['editmovie']
            medit=Movie.query.get(editmovie)
            if medit==None:
                flash("Movie does not exist")
                return render_template("admin.html",movies=allmovie()) 
            
            updateimage(request.files['image'],medit)   #UPDATES IMAGE IF ADMIN UPLOADS ANY            
            title=request.form['title']
            director=request.form['director']
            ryear=request.form['year']
            rating=request.form['rating']
            about=request.form['about']
            duration=request.form['duration']
            actors=request.form['actors']
            if (title=='' or director=='' or ryear=='' or about=='' or rating=='' or duration=='' or actors==''):
                flash('Some fields are empty')
                return render_template("editmovie.html", p=medit)
            elif( 0>float(rating) or float(rating) >10):
                flash('Invalid Imdb')
                return render_template("editmovie.html", p=medit)

            elif(returntime(duration)==time(int(0),int(0),int(0))):
                flash('Invalid duration')        
                return render_template("editmovie.html", p=medit)

            medit.name=title
            medit.director_name=director
            medit.about=about
            medit.release_date=returndate(ryear)  
            medit.imdb_rating=rating
            medit.duration=returntime(duration)
          
            actors=actors.replace("\'","\"")
        
            try:
                b=isinstance(json.loads(actors),list)
                medit.actors=json.loads(actors)
                if(b):
                    try:
                        db.session.commit()
                        return render_template("admin.html", movies=allmovie())
                    except:
                        flash("Error Occured in database")
                        return render_template("editmovie.html", p=medit)
                else:
                    flash('Fill actors list in given format')
                    return render_template("editmovie.html", p=medit)
                    
            except:
                flash('Fill actors list in given format') 
                return render_template("editmovie.html",p=medit)     
            
        elif 'cancel' in request.form:
            return render_template("admin.html", movies=allmovie())
        elif  'delete' in request.form:
            d=request.form['delete'] 
            dm = Movie.query.get(d)
            if dm == None:
                flash("Invalid Movie or Movie already deleted")
                return render_template("admin.html", movies=allmovie())
            try:
                db.session.delete(dm)
                db.session.commit()
            except:
                db.session.rollback()
                flash("Cannot delete due to Venue Constraints")   
                return render_template("admin.html", movies=allmovie())
            return render_template("admin.html", movies=allmovie())
        elif 'addamovie' in request.form:           
            return render_template("newmovie.html")
        elif 'newmovie' in request.form:
            
            image=request.files['image']
            image.save('website/static/new.png')
            title=request.form['title']
            director=request.form['director']
            ryear=request.form['year']
            rating=request.form['rating']
            about=request.form['about']
            duration=request.form['duration']
            actors=request.form['actors']

            if (image.filename=='' or title==None or director=='' or ryear=='' or about=='' or rating=='' or duration=='' or actors==''):
                flash('Some fields are empty')
                return render_template("newmovie.html")
            elif( 0>float(rating) or float(rating) >10):
                flash('Invalid Imdb')
                return render_template("newmovie.html")

            elif(returntime(duration)==time(int(0),int(0))):
                flash('Invalid duration')
            actors=actors.replace("\'","\"")   
            duration=returntime(duration)
  
            try:
                json.loads(actors)
                b=isinstance(json.loads(actors),list)
                if (b):
                    ryear=returndate(ryear)                  
                    with open('website/static/new.png', 'rb') as imgfile:
                        photo= imgfile.read() 
                        
                    new_movie = Movie(name=title,duration=duration,actors=actors,about=about,imdb_rating=rating,release_date=ryear,director_name=director, image=photo)
                    db.session.add(new_movie)
                    db.session.commit()
                    return render_template("admin.html", movies=allmovie())
                else:    
                    flash('Fill actors list in given format')
                    return render_template("newmovie.html")
            except:
                flash('Fill actors list in given format')   
                return render_template("newmovie.html")   
    

@views.route('/admin/venue', methods=['GET', 'POST' ])  # URL endpoint for venue interface where user can edit, add and delete venue
@admin_required                                         # depending upon conditions constraints.             
@login_required
def venue(): 
    from .models import Venue

    today = datetime.today().strftime('%Y-%m-%d')
    fifthday = (datetime.today() + timedelta(days=4)).strftime('%Y-%m-%d')
    if request.method=='GET':
        return render_template("venue.html", venue=allvenue())
    
    if request.method == 'POST' and 'edit' in request.form:
        venedt=request.form['edit']
        ven=Venue.query.get(venedt)
        if(ven==None):
            flash("Venue does not exist")
            return render_template("venue.html", venue=allvenue())
        return render_template("editvenue.html",today=today,theatres=alltheatre(),movies=allmovie(),fifthday=fifthday, ven=ven,seats=seatsbooked(venedt))
    
    if request.method == 'POST' and 'venuesearch' in request.form:
        venedt=request.form.get('editme')
        ven=Venue.query.get(venedt)
        if(ven==None):
            flash("Venue does not exist")
            return render_template("venue.html", venue=allvenue())
        return render_template("editvenue.html",today=today,theatres=alltheatre(),movies=allmovie(),fifthday=fifthday, ven=ven,seats=seatsbooked(venedt))
    
    elif request.method == 'POST' and 'editvenue' in request.form:
        venedt=request.form['editvenue']
        ven=Venue.query.get(venedt)
        movie_id=request.form.get('movie_id')
        theatre_id=request.form.get('theatre_id')
        venue_date=request.form.get('date')
        timing=request.form.get('timing')
        cost=request.form.get('cost')
        if (movie_id=='' or theatre_id=='' or venue_date=='' or timing=='' or cost==''):
            flash('Fill in all the details')
            return render_template("editvenue.html",today=today,fifthday=fifthday,theatres=alltheatre(),movies=allmovie(), ven=ven,seats=seatsbooked(venedt))
        vdate=returndate(venue_date)
        ntiming=returntime(timing)
        if ( returndate(today)>vdate or returndate(fifthday)<vdate):
            flash('Invalid Realise Date')
            return render_template("editvenue.html",today=today,fifthday=fifthday,theatres=alltheatre(),movies=allmovie(), ven=ven,seats=seatsbooked(venedt))
        elif(int(cost)<80):
            flash('Cost Cannot be less than 80')  
            return render_template("editvenue.html",today=today,fifthday=fifthday,theatres=alltheatre(),movies=allmovie(), ven=ven,seats=seatsbooked(venedt))
        
        
        venueexist=Venue.query.filter_by(movie_id=movie_id,theatre_id=theatre_id,date=vdate,time=ntiming,cost=int(cost)).first()
        if venueexist:
            flash("Venue already exist")
            return render_template("venue.html", venue=allvenue())
        ven.movie_id=movie_id
        ven.theatre_id=theatre_id
        ven.date=vdate
        ven.time=ntiming
        ven.cost=cost
        db.session.commit()
        return render_template("venue.html", venue=allvenue())
                
    elif request.method == 'POST' and 'delete' in request.form:
        ventodel=request.form['delete']
        vendel = Venue.query.get(ventodel)
        if vendel==None:
            flash("Venue does not exist or already deleted")
            return render_template("venue.html", venue=allvenue())
        try:
            db.session.delete(vendel)
            db.session.commit()
            return render_template("venue.html", venue=allvenue())
        except:
            db.session.rollback()
            flash("Cannot delete Venue due to Order constraint")
            return render_template("venue.html", venue=allvenue())
    
    elif request.method == 'POST' and 'createvenue' in request.form:        
        return render_template("createvenue.html",movies=allmovie(),theatres=alltheatre(), today=today, fifthday=fifthday)
    
    elif request.method == 'POST' and 'newvenue' in request.form:
        movie_id=request.form.get('movie_id')
        theatre_id=request.form.get('theatre_id')
        venue_date=request.form.get('date')
        timing=request.form.get('timing')
        cost=request.form.get('cost')
        if (movie_id=='' or theatre_id=='' or venue_date=='' or timing=='' or cost==''):
            flash('Fill in all the details')
            return render_template("createvenue.html",movies=allmovie(),theatres=alltheatre(), today=today, fifthday=fifthday)
        vdate=returndate(venue_date)
        ntiming=returntime(timing)
        if ( returndate(today)>vdate or returndate(fifthday)<vdate):
            flash('Invalid Venue Date')
            return render_template("createvenue.html",movies=allmovie(),theatres=alltheatre(), today=today, fifthday=fifthday)
        elif(int(cost)<80):
            flash('Cost Cannot be less than 80')  
            return render_template("createvenue.html",movies=allmovie(),theatres=alltheatre(), today=today, fifthday=fifthday)
        venueexist=Venue.query.filter_by(movie_id=movie_id,theatre_id=theatre_id,date=vdate,time=ntiming,cost=cost).first()
        if venueexist:
            flash('Venue already exist')
            return render_template("venue.html", venue=allvenue())
        newvenue=Venue(movie_id=movie_id,theatre_id=theatre_id,date=vdate,time=ntiming,cost=cost)
        db.session.add(newvenue)
        db.session.commit()
        return render_template("venue.html", venue=allvenue())
        
    elif request.method == 'POST' and 'back' in request.form: 
        return render_template("venue.html", venue=allvenue()) 

@views.route('/theatre', methods=['GET','POST'])  # URL endpoint for theatres where user can edit, add and delete theatres
@admin_required                                   # depending upon conditions and constraints.
@login_required
def theatre(): 
    if request.method=='GET':
        return render_template("theatre.html",theatres=alltheatre())
    
    elif request.method == 'POST' and 'edit' in request.form:
        tid=request.form['edit']
        theatre=Theatre.query.get(tid)
        return render_template("edittheatre.html",theatre=theatre)
    
    elif request.method == 'POST' and 'delete' in request.form:
        tid=request.form['delete']
        theatre = Theatre.query.get(tid)
        if theatre==None:
            flash("Theatre does not exist or already deleted")
            return render_template("theatre.html",theatres=alltheatre())
        try:
            db.session.delete(theatre)
            db.session.commit()
        except:
            db.session.rollback()
            flash("Cannot delete due to Venue constarints")  
            return render_template("theatre.html",theatres=alltheatre())
        return render_template("theatre.html",theatres=alltheatre())

    elif request.method == 'POST' and 'addtheatre' in request.form:
        return render_template("addtheatre.html")
    
    elif request.method == 'POST' and 'cancel' in request.form:
        return render_template("theatre.html",theatres=alltheatre())
    
    elif request.method == 'POST' and 'newtheatre' in request.form:
        theatrename=request.form['theatrename']
        theatrecity=request.form['city']
        capacity=request.form['capacity']
        if (theatrename=='' or theatrecity=='' or capacity.isdigit()==False):
            flash("Fill in all proper details")
            return render_template("addtheatre.html")
        theatreexist=Theatre.query.filter_by(name=theatrename,city=theatrecity,capacity=capacity).first()
        if theatreexist:
            flash("Theatre already exists")
            return render_template("theatre.html",theatres=alltheatre())
        newtheatre=Theatre(name=theatrename,city=theatrecity,capacity=capacity)
        db.session.add(newtheatre)
        db.session.commit()
        return render_template("theatre.html",theatres=alltheatre())
    
    elif request.method == 'POST' and 'edittheatre' in request.form:
        edittheatre=request.form['edittheatre']
        theatrename=request.form['theatrename']
        theatrecity=request.form['city']
        capacity=request.form['capacity']
        editid=Theatre.query.get(edittheatre)
        if (theatrename=='' or theatrecity=='' or capacity.isdigit()==False):
            flash("Fill in all proper details")
            return render_template("addtheatre.html")        
        editid.name=theatrename
        editid.city=theatrecity
        editid.capacity=capacity
        db.session.commit()
        return render_template("theatre.html",theatres=alltheatre())
    
